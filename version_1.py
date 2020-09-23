import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np
from openpyxl import Workbook
# from copy import copy
import copy


# xls = pd.ExcelFile('ا. - لحن اللى التوزيع')
# df1 = pd.read_excel(xls, 'ك مارجرحس حداثق المعادى')
# df2 = pd.read_excel(xls, 'ك مارجرجس بكوتسيكا')

# file_name = "ا. - لحن نى اثنوس تيرو.xlsx"
# xls = pd.read_excel(file_name, sheet_name = "ك مارمرقس بالمعادى")
# xls_1 = pd.read_excel(file_name, sheet_name = "ك الانبا بيشوى بزهراء المعادى")
# print (xls)

# xls.to_excel("output.xlsx", engine='xlsxwriter', mode='a', sheet_name="")  
# xls_1.to_excel("output_1.xlsx")  

# main_file = "ا. - لحن نى اثنوس تيرو.xlsx"

# book = load_workbook(main_file)
# writer = pd.ExcelWriter(main_file, engine='openpyxl') 
# writer.book = book
# sheet_1 = book.worksheets[0]
# print(sheet_1['B10'].value)

# for row in sheet_1.iter_rows(min_row=6, max_col=6, max_row=40, values_only=True):
#     listo = list(row)

#     print (listo[1])
#     print (listo[2])
#     print (listo[3])
#     print (listo[4])
#     print (listo[5])
#     input("e")
        

# print (sheet_1.cell(row=10, column=2))
# dicy = {}
# for ws in book.worksheets:
#     print (ws.title)
#     x = input("do u want this?")
#     if x == 1:
#         dicy[ws.title] = ws
        

# writer.sheets = dicy

# writer.sheets = dict( (ws.title, ws) for ws in book.worksheets)

# data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])
# sh1 =  list(writer.sheets.keys())
# print (sh1[0])
# sh1 = writer.sheets[sh1[0]]
# print (sh1)
# print (sh1.cell(2,3))

# print ("ev")
# writer.save()

# book.save("ev.xlsx")

def reverse(string):
    # strings =(list(string))
    out = ""
    n = len(string)
    for i in range (len(string)):
        out += string[n-1-i]
    return out

def fill_random(file_name):
    '''
        Fill data randomly for validation purposes
    '''
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl') 
    writer.book = book

    for sheet in  book.worksheets:
        print (f'Working with file: {reverse(file_name)} \tsheet: {reverse(sheet.title)}')
        if sheet['A5'].value != "م":
            print ("Error: the sixth row is not the start!!")
        cell_index = 6
        for row in sheet.iter_rows(min_row=6, max_col=6, values_only=True):
            listo = list(row)
            # print (listo)
            #exit when reach the end
            if listo[1] == None:
                break
            
            if listo[0] == None and listo[1] != None:
                print("\nDOUBLE!")
                print (f'WARNING: file name: {reverse(file_name)} \tsheet: {reverse(sheet.title)} \tCell: {cell_index} \tFound a DOUBLE!')
                print (f'name: {reverse(listo[1])}\n\n')
                cell_index += 1
                continue

            pos = 'C'+str(cell_index)
            if sheet[pos].value == None:
            # if  sheet[pos]._style.fillId == 0 :
                rand_int = np.random.randint(0,4)
                sheet['C'+str(cell_index)] = rand_int

                rand_int = np.random.randint(0,3)
                sheet['D'+str(cell_index)] = rand_int

                rand_int = np.random.randint(0,3)
                sheet['E'+str(cell_index)] = rand_int

            cell_index+=1
    writer.save()


def check_values_bound(file_name=None):
    '''
        For each sheet we have, check the bounded values
        if they exceed (4,3,3), then print out the file name/sheet name/ and the row
    '''
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl') 
    writer.book = book

    for sheet in book.worksheets:
        print (f'Working with sheet: {(sheet.title)}')
        if sheet['A5'].value != "م":
            print ("Error: the sixth row is not the start!!")

        skip_5_rows = 5
        # dst_sheet = dst_book[src_sheet.title]
        for row_src in sheet.rows:
            if skip_5_rows > 0:
                skip_5_rows -= 1
                continue
            if row_src[2].value != None:
                try:
                    if row_src[2].value > 4 or row_src[3].value > 3 or row_src[4].value > 3\
                        or row_src[2].value <0 or row_src[3].value <0 or row_src[4].value <0 :
                        print (f'HUGE ERROR: file name: {reverse(file_name)} \tsheet: {reverse(sheet.title)} \trow_index: {row_src[0].value} Validation error!')
                        input("Ev") 
                except:
                    print (f'row index: {row_src[0].value}')

    
    print ("All is fine")
                        

def merge_separated (source_1, source_2):
    wb = Workbook()
    
    book = load_workbook(source_1)
    writer = pd.ExcelWriter(source_1, engine='openpyxl') 
    writer.book = book

    dicy = {}
    for sheet in book.worksheets:
        # dicy[sheet.title] = sheet
        sh = wb.create_sheet(sheet.title)
        # for row_src in sheet.iter_rows(min_row=1, max_col=6, values_only=True):
        for row_src in sheet.rows:
            # listo = list(row_src)
            # if listo[1] == None and listo[0] == None and listo[2] == None and listo[3] == None and listo[4] == None:
            #     break

            for cell in row_src:
                if cell == None:
                    continue
                try:
                    new_cell = sh.cell(row=cell.row, column=cell.col_idx,
                            value= cell.value)
                    if cell.has_style:
                        new_cell.font = copy.copy(cell.font)
                        new_cell.border = copy.copy(cell.border)
                        new_cell.fill = copy.copy(cell.fill)
                        new_cell.number_format = copy.copy(cell.number_format)
                        new_cell.protection = copy.copy(cell.protection)
                        new_cell.alignment = copy.copy(cell.alignment)
                except AttributeError:
                    print ('sorry')
        destination = source_1.split('.')[1] + " + " +source_2.split('.')[1] + ".xlsx"

        wb.save(destination)
        return


        # max_rows = 1
        # for row_src in sheet.iter_rows(min_row=6, max_col=6, values_only=True):
        #     listo = list(row_src)
        #     if listo[1] == None and  listo[0] == None :
        #         break
        #     max_rows += 1

        # for i in range (1,max_rows+7):
        #     sh['A'+str(i)] = sheet['A'+str(i)].value
        #     sh['B'+str(i)] = sheet['B'+str(i)].value
        #     sh['C'+str(i)] = sheet['C'+str(i)].value
        #     sh['D'+str(i)] = sheet['D'+str(i)].value
        #     sh['E'+str(i)] = sheet['E'+str(i)].value
        #     sh['F'+str(i)] = sheet['F'+str(i)].value



    book = load_workbook(source_2)
    writer = pd.ExcelWriter(source_2, engine='openpyxl') 
    writer.book = book

    for sheet in  book.worksheets:
        # dicy[sheet.title] = sheet
        sh = wb.create_sheet(sheet.title)
        max_rows = 1
        for row_src in sheet.iter_rows(min_row=6, max_col=6, values_only=True):
            listo = list(row_src)
            if listo[1] == None and  listo[0] == None :
                break
            max_rows += 1

        for i in range (1,max_rows+1+7):
            sh['A'+str(i)] = sheet['A'+str(i)].value            
            sh['B'+str(i)] = sheet['B'+str(i)].value
            sh['C'+str(i)] = sheet['C'+str(i)].value
            sh['D'+str(i)] = sheet['D'+str(i)].value
            sh['E'+str(i)] = sheet['E'+str(i)].value
            sh['F'+str(i)] = sheet['F'+str(i)].value

    destination = source_1.split('.')[1] + " + " +source_2.split('.')[1] + ".xlsx"
    
    wb.save(destination)
    return destination

def print_info (file_name, sheet_name):
    print (f'Working with file: \t{file_name} \tin sheet: \t{sheet_name}.')

def fill_main(main_file, source_1, mian_extra = True):
    print (f'Destination file: \t{main_file}')
    print (f'Source file: \t\t{source_1}')

    src_book = load_workbook(source_1)
    src_writer = pd.ExcelWriter(source_1, engine='openpyxl') 
    src_writer.book = src_book

    dst_book = load_workbook(main_file)
    dst_writer = pd.ExcelWriter(main_file, engine='openpyxl') 
    dst_writer.book = dst_book
    dst_sheets = list(dst_book.sheetnames)
    dst_sheets_dic = {}
    for src_sheet in src_book.worksheets:
        print_info (source_1, src_sheet.title)

        # find the same sheet in the main file!
        # if it does not exist, print an error message!
        if src_sheet.title in dst_sheets:
            skip_5_rows = 5
            dst_sheet = dst_book[src_sheet.title]
            for row_src in src_sheet.rows:
                if skip_5_rows > 0:
                    skip_5_rows -= 1
                    continue

                for i, cell in enumerate(row_src):
                    if cell == None:
                        continue

                    if i >= 5:
                        break
                    try:
                        if mian_extra:
                            new_cell = dst_sheet.cell(row=cell.row, column=cell.col_idx, value= cell.value)
                            
                        else:
                            if cell.col_idx == 1 or cell.col_idx == 2:
                                continue
                            new_cell = dst_sheet.cell(row=cell.row, column=cell.col_idx+4, value= cell.value)
                            
                        if cell.has_style:
                            new_cell.font = copy.copy(cell.font)
                            new_cell.border = copy.copy(cell.border)
                            new_cell.fill = copy.copy(cell.fill)
                            new_cell.number_format = copy.copy(cell.number_format)
                            new_cell.protection = copy.copy(cell.protection)
                            new_cell.alignment = copy.copy(cell.alignment)
                    except AttributeError:
                        # print ('sorry allignment error occured!')
                        print("\nDOUBLE!")
                        print (f'WARNING: Source File: {reverse(source_1)} \tsheet: {reverse(src_sheet.title)} \tRow: {cell.row} \tFound a DOUBLE!\n\n')
                        break
            dst_sheets_dic [dst_sheet.title] = dst_sheet
        else:
            print (f'Error: sheet \t{src_sheet.title} \tdoes not exist in file: \t{main_file}')
            input ("Evram!!")
            dst_writer.sheets = dst_sheets_dic
            dst_writer.save()
            return

    dst_writer.sheets = dst_sheets_dic
    dst_writer.save()

    
def add_avg(file_path):
    '''
        Add the avg grades for a sheet
    '''
    src_book = load_workbook(file_path)
    src_writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    src_writer.book = src_book
    
    dst_sheets_dic = {}
    for src_sheet in src_book.worksheets:
        skip_5_rows = 5
        sevens = []
        threes = []
        # last_idx = None
        # empty_flag = False

        for row_src in src_sheet.rows:
            if skip_5_rows > 0:
                skip_5_rows -= 1
                continue
            
            # if row_src[0].value == None or row_src[1].value == None:
            #     empty_flag = True
            #     break
            
            # if last_idx == row_src[0]:
            #     continue

            # last_idx = row_src[0].value

            # sevens.append(row_src[5].value)
            # threes.append(row_src[10].value)
            if row_src[2].value != None and row_src[3].value != None and row_src[4].value != None:
                sevens.append((row_src[2].value+row_src[3].value+row_src[4].value)*0.7)
            if row_src[5].value != None and row_src[6].value != None and row_src[7].value != None:
                threes.append((row_src[6].value+row_src[7].value+row_src[8].value)*0.3)


        # if not empty_flag:
        if len(sevens) > 0:
            src_sheet['F4'] = np.average(np.array(sevens))
        if len(threes) > 0:
            src_sheet['J4'] = np.average(np.array(threes))
        dst_sheets_dic [src_sheet.title] = src_sheet

    src_writer.sheets = dst_sheets_dic
    src_writer.save()


def add_fraction(file_path):
    src_book = load_workbook(file_path)
    src_writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    src_writer.book = src_book
    dst_sheets_dic = {}

    for src_sheet in src_book.worksheets:
        seven_avg = src_sheet['F4']
        three_avg = src_sheet['J4']

        #If sheet is empty
        if seven_avg.value == None and three_avg.value == None:
            continue
        
        if three_avg.value == None:
            src_sheet['K4'] = seven_avg.value
            summ = seven_avg.value
        else:
            src_sheet['K4'] = seven_avg.value + three_avg.value
            summ = seven_avg.value + three_avg.value
        #count them
        skip_5_rows = 5
        count = 0
        for row_src in src_sheet.rows:
            if skip_5_rows > 0:
                skip_5_rows -= 1
                continue

            if row_src[0].value == None and row_src[1].value == None:
                break

            if row_src[0].value != None:  
                count += 1
        
        src_sheet['L2'] = "Count"
        src_sheet['M2'] = count

        Frac = 1
        src_sheet['L3'] = "Frac"
        if count >= 5 and count <= 10:
            Frac = 1.01

        if count > 10 and count <= 15:
            Frac = 1.02

        if count > 15 and count <= 20:
            Frac = 1.03

        if count > 20 and count <= 25:
            Frac = 1.04

        if count > 25 and count <= 30:
            Frac = 1.05
            
        if count > 30:
            Frac = 1.06
        
        src_sheet['M3'] = Frac

        src_sheet['L4'] = "Final"
        src_sheet['M4'] = Frac * (summ)
        
        print(f'Count: {count} \tFrac: {Frac} \tSum Value: {summ} \tFinal: {Frac * (summ)}')
        input("GOOD?")

        dst_sheets_dic [src_sheet.title] = src_sheet

    src_writer.sheets = dst_sheets_dic
    src_writer.save()

def get_tops (file_path):
    src_book = load_workbook(file_path)
    src_writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    src_writer.book = src_book
    data = []
    for src_sheet in src_book.worksheets:
        count = 0
        #If sheet is empty
        if src_sheet['M2'].value == None:
            print (f'Sheet: {reverse(src_sheet.title)} is EMP')
            continue

        skip_5_rows = 5
        element = []
        for row_src in src_sheet.rows:
            if skip_5_rows > 0:
                skip_5_rows -= 1
                continue
            
            if row_src[0].value == None and row_src[1].value == None:
                break
            
            
            if row_src[0].value == None and row_src[1].value != None:
                new_element = copy.deepcopy(element)
                new_element[1] = row_src[1].value + "DOUBLE"
                data.append(new_element)
                
                continue

            element = []
            points = ((row_src[2].value+row_src[3].value+row_src[4].value)*0.7)
            if row_src[8].value != None:
                points += ((row_src[8].value+row_src[6].value+row_src[7].value)*0.3)

            element.append (points) #points
            element.append (row_src[1].value) #name
            element.append (src_sheet.title) #church
            data.append(element)
            count += 1
            

        print (f'From: {reverse(src_sheet.title)} \tread {count}')
    data.sort(reverse=True, key= lambda x: x[0])
    print (data[0])
    input("ev")

def get_total (file_path):
    src_book = load_workbook(file_path)
    src_writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    src_writer.book = src_book
    count = 0
    for src_sheet in src_book.worksheets:
        if src_sheet['M2'].value != None:
            count += src_sheet['M2'].value
    return count
    

if __name__ == "__main__":
    
    # file_path = "dst/" + "جروب A - استمارة تقييم مهرجان 2020 - المستوى الثالث.xlsx"
    # add_fraction(file_path)
    # input("stop")
    
    # file_path = "dst/" + 'جروب A - استمارة تقييم مهرجان 2020 - المستوى الثالث.xlsx'
    # get_tops(file_path)
    # input("stop")
    
    files = os.listdir("dst/")
    total = 0
    for folder_ in files:
        file_path = "dst/" + folder_
        total += get_total(file_path)
    print (total)
    input("e")
    if os.listdir("dst/") == []:
        print (f'Error: no destination files!')
    
    if os.listdir("src_main/") == []:
        print (f'Error: no mian source files!')
    
    if os.listdir("src_extra/") == []:
        print (f'Error: no extra files!')
        
    dst_path = "dst/" + os.listdir("dst/")[0]
    input("Should I start working?")
    for src_folders in ["src_main/", "src_extra/"]:
        srcs = os.listdir(src_folders)
        if src_folders.split("_")[1] == "main/":
            mian_extra = True
        else:
            mian_extra = False
        for src in srcs:
            src_path = src_folders + src
            check_values_bound(src_path)
            fill_main(dst_path, src_path, mian_extra=mian_extra)

    print("Done, will add the avg")
    input("OK?")
    file_at_dst = dst_path
    add_avg(file_at_dst)
    
    

    # dest = "استمارة تقييم مهرجان 2020.xlsx"
    # source_1 = "ا.ابونا اغسطينوس كامل - لحن تين او اوشت (( او )) طاى شورى.xlsx"
    # source_2 = "ا.ابرام عادل - لحن تين او اوشت (( او )) طاى شورى.xlsx"
    # source_3 = "ا.بيشوى عادل - لحن اللى التوزيع.xlsx"

    # for file_path in os.listdir("fill_these/"):
    #     main_file = "fill_these/" + file_path
    #     fill_random(main_file)
    # check_values_bound(main_file)
    # merged = merge_separated (source_1, source_2)

    # fill_main(dest, source_1)

    